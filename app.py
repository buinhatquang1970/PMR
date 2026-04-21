import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import io
import os
import html
import logging
from logging.handlers import RotatingFileHandler
import uuid
import time
from datetime import datetime, timedelta, timezone
from tool_tinh_toan import ToolAnDinhTanSo
import importlib
import gc  # --- BỔ SUNG: THƯ VIỆN GIẢI PHÓNG BỘ NHỚ CHỦ ĐỘNG ---

ACCESS_LOG_FILE = 'total_access.txt'

def get_total_access():
    """Đọc tổng lượt truy cập từ file"""
    if not os.path.exists(ACCESS_LOG_FILE):
        return 0
    with open(ACCESS_LOG_FILE, 'r') as f:
        try:
            return int(f.read().strip())
        except:
            return 0

def increment_total_access():
    """Tăng tổng lượt truy cập và lưu vào file"""
    count = get_total_access() + 1
    with open(ACCESS_LOG_FILE, 'w') as f:
        f.write(str(count))
    return count

@st.cache_resource
def get_counted_sessions():
    """Bộ nhớ tạm để lưu danh sách session_id đã được đếm trong phiên chạy này"""
    return set()

# --- IMPORT AN TOÀN CHO BIẾN MÀU SẮC & QUY HOẠCH TẦN SỐ ---
try:
    import config
    importlib.reload(config)
    PRIORITY_HIGHLIGHT_COLOR = getattr(config, 'PRIORITY_HIGHLIGHT_COLOR', '#F6BE00')
    ALLOC_VHF = getattr(config, 'FREQUENCY_ALLOCATION_VHF', [])
    ALLOC_UHF = getattr(config, 'FREQUENCY_ALLOCATION_UHF', [])
except:
    PRIORITY_HIGHLIGHT_COLOR = '#F6BE00'
    ALLOC_VHF = []
    ALLOC_UHF = []

# =============================================================================
# CẤU HÌNH HỆ THỐNG GHI LOG (LOGGING SETUP) - ĐÃ FIX MÚI GIỜ GMT+7
# =============================================================================
LOG_FILE = 'pmr_tool_usage.log'
ADMIN_PASSWORD = '123456' # Mật khẩu Admin

# --- HÀM LẤY IP NGƯỜI DÙNG ---
def get_remote_ip():
    """Lấy IP thật của người dùng sử dụng st.context.headers"""
    try:
        if hasattr(st, "context") and st.context.headers:
            headers = st.context.headers
            ip_headers = [
                "X-Forwarded-For", "X-Real-Ip", "Forwarded", "X-Client-Ip", "Remote-Addr"
            ]
            for key in ip_headers:
                if key in headers:
                    val = headers[key]
                    if val:
                        return val.split(',')[0].strip()
            return headers.get("Host", "Unknown_Host")
    except Exception:
        pass
    return "127.0.0.1"

def setup_logging():
    logger = logging.getLogger("PMR_Tool_Logger")
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        
        # --- [QUAN TRỌNG] TẠO CONVERTER MÚI GIỜ VIỆT NAM (GMT+7) ---
        def nam_time(*args):
            # Lấy giờ UTC hiện tại và cộng thêm 7 giờ
            utc_now = datetime.now(timezone.utc)
            vn_now = utc_now.astimezone(timezone(timedelta(hours=7)))
            return vn_now.timetuple()

        # Định dạng log
        formatter = logging.Formatter('%(asctime)s - [IP:%(client_ip)s] - [%(levelname)s] - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        
        # Gán hàm chuyển đổi giờ VN cho formatter
        formatter.converter = nam_time 
        
        # 1. Ghi ra file
        file_handler = RotatingFileHandler(LOG_FILE, maxBytes=5*1024*1024, backupCount=3, encoding='utf-8')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        
        # 2. Ghi ra console
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)
    return logger

logger = setup_logging()

# Hàm wrapper để ghi log kèm IP
def log_info(msg):
    ip = get_remote_ip()
    logger.info(msg, extra={'client_ip': ip})

def log_warning(msg):
    ip = get_remote_ip()
    logger.warning(msg, extra={'client_ip': ip})

def log_exception(msg):
    ip = get_remote_ip()
    logger.exception(msg, extra={'client_ip': ip})

# --- APP VERSION ---
try:
    file_timestamp = os.path.getmtime(__file__)
    vn_tz = timezone(timedelta(hours=7))
    dt_vn = datetime.fromtimestamp(file_timestamp, vn_tz)
   # Lấy giờ theo định dạng 24h (0-23) thay vì 12h
    APP_VERSION = f"v{dt_vn.strftime('%d%m%y')}.{dt_vn.hour}"
except Exception:
    APP_VERSION = "v300126.14"


# =============================================================================
# TỐI ƯU HÓA BỘ NHỚ (RAM) - MỤC 2
# =============================================================================
@st.cache_resource(show_spinner="Đang nạp file vào bộ nhớ, vui lòng chờ...")
def get_tool_instance(uploaded_files):
    """Giữ đối tượng Tool trong bộ nhớ (Cache) để tránh đọc lại file nhiều lần tốn RAM"""
    if uploaded_files:
        try:
            instance = ToolAnDinhTanSo(uploaded_files)
            gc.collect() # Dọn RAM tức thì sau khi đọc
            return instance
        except Exception as e:
            raise e
    return None
# =============================================================================
# =============================================================================
# THEO DÕI SỐ LƯỢNG NGƯỜI DÙNG ONLINE
# =============================================================================
@st.cache_resource
def get_active_users_dict():
    """Tạo một từ điển dùng chung để lưu thời gian hoạt động của tất cả người dùng"""
    return {}

# --- CẤU HÌNH TRANG ---
st.set_page_config(page_title=f"PMR tool ({APP_VERSION})", layout="wide")

# --- CSS TÙY CHỈNH (Dùng chung) ---
st.markdown("""
    <style>
        .block-container { padding-top: 1rem !important; padding-bottom: 2rem; }
        header[data-testid="stHeader"] {
            height: 2rem !important; 
            background-color: transparent !important;
        }
        h2 { font-size: 1.3rem !important; margin-top: 0.5rem; margin-bottom: 0.2rem !important; }
        h3 { font-size: 0.95rem !important; padding-top: 0.2rem !important; padding-bottom: 0.2rem !important; }
        div[data-testid="stMarkdownContainer"] > p { margin-bottom: -3px !important; font-weight: 500; }
        [data-testid="stHorizontalBlock"] { gap: 0.1rem !important; }
        .stCaption { font-size: 0.7rem; margin-top: -5px; color: #555; }
        hr { margin-top: 0.5rem !important; margin-bottom: 0.5rem !important; }
        
        /* ===== FILE UPLOADER - GIAO DIỆN HOÀN CHỈNH ===== */
        [data-testid='stFileUploader'] { margin-bottom: -30px !important; }
        [data-testid='stFileUploader'] section { padding: 0.5rem !important; min-height: 0px !important; }
        
        /* 1. XÓA SẠCH rác tiếng Anh (Drag drop, Limit 200MB...) */
        [data-testid='stFileUploader'] section > div > div:not(:has(button)) {
            display: none !important;
        }
        [data-testid='stFileUploader'] section > div > small,
        [data-testid='stFileUploader'] section > div > span,
        [data-testid='stFileUploader'] section > div > p {
            display: none !important;
        }

        /* 2. CHÈN HƯỚNG DẪN 3 BƯỚC (Căn lề trái thẳng tắp) */
        [data-testid='stFileUploader'] section > div::before {
            content: "1. Xuất dữ liệu mới nhất từ PM cấp phép \\A 2. Lưu dưới dạng Excel Workbook (.xlsx) \\A 3. Bấm Chọn Files để nạp (Tối đa 2 files)";
            white-space: pre-wrap !important;
            display: block !important;
            font-weight: bold !important;
            color: #333 !important;
            text-align: left !important;
            font-size: 14px !important;
            line-height: 1.6 !important;
            width: 100% !important;
            margin: 0 0 10px 0 !important;
        }

        /* 3. VIỆT HÓA NÚT BẤM (Xóa Browse files -> Đổi thành Chọn File Excel) */
        /* Bước 3.1: Tàng hình hoàn toàn chữ tiếng Anh cũ */
        [data-testid='stFileUploader'] section button, 
        [data-testid='stFileUploader'] section button * {
            font-size: 0px !important; 
            color: transparent !important;
        }
        /* Bước 3.2: Bơm chữ tiếng Việt vào giữa nút */
        [data-testid='stFileUploader'] section button::after {
            content: '📂 CHỌN FILES' !important;
            font-size: 15px !important;
            color: #31333F !important;
            display: block !important;
            visibility: visible !important;
            font-weight: 600 !important;
        }
        /* ===== STYLE CHO FILE ĐÃ TẢI LÊN ===== */
        [data-testid="stUploadedFile"] {
            display: flex !important;
            align-items: center !important;
            justify-content: space-between !important;
            gap: 12px !important;
            padding: 10px 12px !important;
            background-color: #f9f9f9 !important;
            border: 1px solid #e0e0e0 !important;
            border-radius: 6px !important;
            margin-bottom: 8px !important;
            margin-top: 10px !important;
            font-size: 14px !important;
        }

        /* Hiển thị tên file - sửa font-size */
        [data-testid="stUploadedFile"] span {
            font-size: 14px !important;
            display: inline !important;
            color: #333 !important;
        }

        /* Icon file trước tên */
        [data-testid="stUploadedFile"]::before {
            content: "📄 " !important;
            margin-right: 8px !important;
            font-size: 1.1rem !important;
        }

        /* ===== STYLE NÚT DELETE (X) ===== */
        [data-testid="stUploadedFile"] button {
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
            min-width: 28px !important;
            min-height: 28px !important;
            width: 28px !important;
            height: 28px !important;
            padding: 0 !important;
            background-color: #ffebee !important;
            border: 1.5px solid #d93025 !important;
            border-radius: 4px !important;
            cursor: pointer !important;
            transition: all 0.2s ease !important;
            flex-shrink: 0 !important;
        }

        [data-testid="stUploadedFile"] button:hover {
            background-color: #ffcdd2 !important;
            border-color: #c5221f !important;
            transform: scale(1.05) !important;
        }

        [data-testid="stUploadedFile"] button:active {
            background-color: #ef5350 !important;
            transform: scale(0.95) !important;
        }

        [data-testid="stUploadedFile"] button span {
            display: inline-flex !important;
            color: #d93025 !important;
            font-weight: bold !important;
            font-size: 1.3rem !important;
            line-height: 1 !important;
        }

        div[data-testid="stColumn"] button[kind="secondary"] { color: #d93025 !important; font-weight: bold !important; border: 1px solid #ddd !important; background-color: #fff !important; width: 100%; transition: all 0.3s; }
        div[data-testid="stColumn"] button[kind="secondary"]:hover { background-color: #fce8e6 !important; border-color: #d93025 !important; color: #d93025 !important; }
#       button[kind="primary"] { font-weight: bold !important; margin-top: 5px; }
        /* ===== ĐỔI MÀU NÚT CHÍNH (TÍNH TẦN SỐ) SANG XANH NƯỚC BIỂN ===== */
        button[kind="primary"] { 
            background-color: #0068C9 !important; 
            border-color: #0068C9 !important;
            color: white !important;
            font-weight: bold !important; 
            margin-top: 5px !important; 
            transition: all 0.3s ease !important;
        }
        /* Hiệu ứng khi di chuột vào (đổi sang xanh đậm hơn một chút) */
        button[kind="primary"]:hover { 
            background-color: #0052a3 !important; 
            border-color: #0052a3 !important; 
        }
        /* Hiệu ứng khi click chuột (đổi sang xanh đậm nhất) */
        button[kind="primary"]:active {
            background-color: #003d7a !important;
            border-color: #003d7a !important;
        }        
        div[data-testid="stTable"] table { width: 100% !important; }
        div[data-testid="stTable"] th { background-color: #f0f2f6 !important; color: #31333F !important; font-size: 1.2rem !important; font-weight: 800 !important; text-align: center !important; white-space: nowrap !important; padding: 15px !important; }
        div[data-testid="stTable"] td { font-size: 1.1rem !important; text-align: center !important; vertical-align: middle !important; padding: 12px !important; min-width: 200px !important; }
        
        div[role="dialog"] { width: 50vw !important; max-width: 50vw !important; left: auto !important; right: 0 !important; top: 0 !important; bottom: 0 !important; height: 100vh !important; margin: 0 !important; border-radius: 0 !important; transform: none !important; display: flex; flex-direction: column; }
        
        div[role="dialog"][aria-label*="Decimal"] {
            width: 30vw !important;             
            min-width: 400px !important;        
            height: max-content !important;     
            left: 50% !important;               
            right: auto !important;
            top: 50% !important;                
            bottom: auto !important;
            transform: translate(-50%, -50%) !important; 
            border-radius: 15px !important;     
            padding-bottom: 20px !important;
        }

        div[data-testid="stSelectbox"] > div, div[data-testid="stSelectbox"] button, div[data-testid="stSelectbox"] select { min-width: 60px !important; max-width: 100% !important; white-space: nowrap !important; overflow: hidden !important; text-overflow: ellipsis !important; display: inline-block !important; }
        .stTextInput, .stSelectbox, .stNumberInput, .stDateInput { min-width: 50px !important; }

        .tooltip-container {
          position: relative;
          display: inline-block;
          cursor: help;
          color: #0068C9;
          font-weight: bold;
          margin-right: 15px;
          z-index: 9999;
        }
        .tooltip-container .tooltiptext {
          visibility: hidden;
          width: 500px;
          background-color: #ffffff;
          color: #333;
          text-align: left;
          border-radius: 6px;
          padding: 15px;
          position: absolute;
          z-index: 10000;
          top: 100%;
          right: 0;
          box-shadow: 0px 8px 16px 0px rgba(0,0,0,0.3);
          border: 1px solid #ddd;
          font-size: 0.9rem;
          font-weight: normal;
          line-height: 1.5;
          max-height: 60vh;
          overflow-y: auto;
        }
        .tooltip-container:hover .tooltiptext {
          visibility: visible;
        }
        .tooltiptext strong { color: #0068C9; }
        
        /* ===== SỬA LỖI XUẤT HIỆN KÉP - KHÔI PHỤC NÚT X (SỨC MẠNH TỐI ĐA) ===== */
        [data-testid="stFileUploader"] [data-testid="stUploadedFile"] button::after {
            content: none !important;
            display: none !important;
        }
        [data-testid="stFileUploader"] [data-testid="stUploadedFile"] button,
        [data-testid="stFileUploader"] [data-testid="stUploadedFile"] button * {
            font-size: 1.5rem !important;
            color: #d93025 !important;
            background-color: transparent !important;
            visibility: visible !important;
        }
    </style>
""", unsafe_allow_html=True)

# --- CÁC HÀM HỖ TRỢ ---
def dms_to_decimal(d, m, s): return d + (m / 60.0) + (s / 3600.0)

# =========================================================================
# HÀM CONVERT ĐƯỢC NÂNG CẤP XỬ LÝ SỐ THẬP PHÂN TUẦN HOÀN (.333333)
# =========================================================================
def decimal_to_dms(decimal):
    sign = -1 if decimal < 0 else 1
    decimal = abs(decimal)
    
    # Tính tổng số giây và làm tròn tới 2 chữ số thập phân
    total_seconds = round(decimal * 3600, 2)
    
    # Bắt các trường hợp lỗi số thập phân tuần hoàn bị cắt cụt 
    # (như .33333 hoặc .66666) khiến tổng giây bị hụt một chút (.98, .99). 
    # Ép làm tròn lên số nguyên để khắc phục.
    fractional_part = total_seconds - int(total_seconds)
    if fractional_part >= 0.98:
        total_seconds = round(total_seconds)
        
    d = int(total_seconds // 3600)
    m = int((total_seconds % 3600) // 60)
    s = round(total_seconds % 60, 2)
    
    return d * sign, m, s

def neutralize_excel_value(val):
    if pd.isna(val): return val
    s = str(val)
    if s and s[0] in ('=', '+', '-', '@'): return "'" + s
    return s
def neutralize_df_for_excel(df):
    """Xử lý rác Excel, tương thích cả Pandas cũ (applymap) và mới (map)"""
    try:
        # Kiểm tra nếu Pandas là bản mới (có hàm map)
        if hasattr(df, 'map'):
            try: 
                return df.map(neutralize_excel_value)
            except Exception: 
                return df.astype(str).map(neutralize_excel_value)
        # Nếu là Pandas bản cũ
        else:
            try: 
                return df.applymap(neutralize_excel_value)
            except Exception: 
                return df.astype(str).applymap(neutralize_excel_value)
    except Exception:
        return df # Fallback an toàn nếu có lỗi bất ngờ
def to_excel(df_input, df_result):
    output = io.BytesIO()
    if df_input is not None: df_input_safe = neutralize_df_for_excel(df_input.copy())
    else: df_input_safe = None
    if 'is_priority' in df_result.columns: df_result_clean = df_result.drop(columns=['is_priority'])
    else: df_result_clean = df_result
    df_result_safe = neutralize_df_for_excel(df_result_clean.copy())
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = 'KET_QUA_TINH_TOAN'
        start_row_result = 1
        if df_input_safe is not None:
            df_input_safe.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)
            start_row_result = len(df_input_safe) + 5
        df_result_safe.to_excel(writer, sheet_name=sheet_name, startrow=start_row_result, index=False)
        worksheet = writer.sheets[sheet_name]
        if df_input_safe is not None:
            worksheet.cell(row=1, column=1, value="I. THÔNG SỐ ĐẦU VÀO").font = openpyxl.styles.Font(bold=True, size=11) if 'openpyxl' in locals() else None
            worksheet.cell(row=start_row_result, column=1, value="II. KẾT QUẢ TÍNH TOÁN").font = openpyxl.styles.Font(bold=True, size=11) if 'openpyxl' in locals() else None
        else:
            worksheet.cell(row=start_row_result, column=1, value="DANH SÁCH KẾT QUẢ")
    return output.getvalue()

@st.dialog("Vị trí trên Google Maps")
def show_map_popup(lat, lon):
    map_url = f"https://www.google.com/maps?q={lat},{lon}&z=15&output=embed"
    st.write(f"📍 Tọa độ: {lat:.5f}, {lon:.5f}")
    components.iframe(map_url, height=600)

# --- HÀM CALLBACK CHUYỂN ĐỔI TỌA ĐỘ ---
def handle_conversion():
    # Lấy giá trị thập phân người dùng vừa nhập trong Pop-up
    dec_lon = st.session_state.get("pop_lon", 0.0)
    dec_lat = st.session_state.get("pop_lat", 0.0)
    
    if dec_lat != 0.0 or dec_lon != 0.0:
        lat_d, lat_m, lat_s = decimal_to_dms(dec_lat)
        lon_d, lon_m, lon_s = decimal_to_dms(dec_lon)
        
        # Cập nhật thẳng vào Session State TRƯỚC khi màn hình được vẽ lại
        st.session_state.lat_d = lat_d
        st.session_state.lat_m = lat_m
        st.session_state.lat_s = lat_s
        st.session_state.lon_d = lon_d
        st.session_state.lon_m = lon_m
        st.session_state.lon_s = lon_s

# --- POP-UP CONVERT ---
@st.dialog("🔄 Nhập Tọa Độ Thập Phân (Decimal -> DMS)")
def show_convert_popup():
    st.markdown("Nhập tọa độ dạng thập phân để quy đổi tự động sang Độ, Phút, Giây:")
    c1, c2 = st.columns(2)
    with c1:
        st.number_input("Kinh độ (Longitude)", value=0.0, format="%.6f", key="pop_lon")
    with c2:
        st.number_input("Vĩ độ (Latitude)", value=0.0, format="%.6f", key="pop_lat")
        
    # Gắn hàm callback vào nút bấm (on_click=handle_conversion)
    if st.button("🧮 CHUYỂN ĐỔI & ÁP DỤNG", type="primary", use_container_width=True, on_click=handle_conversion):
        st.rerun() # Tự động đóng popup và tải lại giao diện

try:
    import openpyxl
except:
    pass

# --- XỬ LÝ KHỞI TẠO SESSION & ĐẾM TRUY CẬP ---
if 'session_id' not in st.session_state:
    st.session_state.session_id = str(uuid.uuid4())[:8] 
    log_info(f"NEW SESSION STARTED | ID: {st.session_state.session_id} | App Version: {APP_VERSION}")
    
    # --- [MỚI] KIỂM TRA CHẶN F5 VÀ ĐẾM LƯỢT TRUY CẬP TỔNG ---
    counted_sessions = get_counted_sessions()
    if st.session_state.session_id not in counted_sessions:
        increment_total_access() # Cộng vào file tổng
        counted_sessions.add(st.session_state.session_id) # Đánh dấu đã đếm

# --- CÁC BIẾN SESSION MẶC ĐỊNH (GIỮ NGUYÊN) ---
if 'results' not in st.session_state: st.session_state.results = None
if 'input_snapshot' not in st.session_state: st.session_state.input_snapshot = None
if 'last_uploaded_file_id' not in st.session_state: st.session_state.last_uploaded_file_id = None
if 'check_result' not in st.session_state: st.session_state.check_result = None
if 'bad_freq_results' not in st.session_state: st.session_state.bad_freq_results = None
if 'active_view' not in st.session_state: st.session_state.active_view = None
if 'admin_logged_in' not in st.session_state: st.session_state.admin_logged_in = False
if 'auto_refresh' not in st.session_state: st.session_state.auto_refresh = False

# --- [MỚI] CẬP NHẬT TRẠNG THÁI ONLINE CỦA NGƯỜI DÙNG ---
active_users = get_active_users_dict()
active_users[st.session_state.session_id] = time.time()



# --- Khởi tạo giá trị mặc định cho Tọa độ để tránh lỗi Widget Value ---
if 'lon_d' not in st.session_state: st.session_state.lon_d = 105
if 'lon_m' not in st.session_state: st.session_state.lon_m = 0
if 'lon_s' not in st.session_state: st.session_state.lon_s = 0.0
if 'lat_d' not in st.session_state: st.session_state.lat_d = 21
if 'lat_m' not in st.session_state: st.session_state.lat_m = 0
if 'lat_s' not in st.session_state: st.session_state.lat_s = 0.0

# --- CẬP NHẬT TRẠNG THÁI ONLINE CỦA NGƯỜI DÙNG ---
active_users = get_active_users_dict()
# Ghi nhận thời điểm hiện tại (bằng giây) mà người dùng này vừa thao tác
active_users[st.session_state.session_id] = time.time()

# =========================================================================
# PHÂN LUỒNG: KIỂM TRA QUERY PARAMS ĐỂ XÁC ĐỊNH GIAO DIỆN ADMIN
# =========================================================================

query_params = st.query_params
is_admin_route = query_params.get("view") == "admin"

if is_admin_route:
    # ---------------------------------------------------------
    # GIAO DIỆN DÀNH RIÊNG CHO ADMIN
    # ---------------------------------------------------------
    st.title("🔐 Hệ thống Quản trị (Admin System)")
    
    if not st.session_state.admin_logged_in:
        st.info("Bạn đang truy cập trang dành riêng cho Quản trị viên.")
        
        # --- KHUNG ĐĂNG NHẬP NHỎ GỌN (TỶ LỆ 3:2:3) ---
        c_login1, c_login2, c_login3 = st.columns([3, 2, 3]) 
        with c_login2:
            with st.form("admin_login_form"):
                st.subheader("Đăng nhập")
                pwd = st.text_input("Mật khẩu:", type="password")
                submitted = st.form_submit_button("Truy cập")
                
                if submitted:
                    if pwd == ADMIN_PASSWORD:
                        st.session_state.admin_logged_in = True
                        st.success("Đăng nhập thành công!")
                        st.rerun()
                    else:
                        st.error("Mật khẩu không đúng!")
    else:
    
# Khi đã đăng nhập thành công
        st.success(f"Xin chào Admin! (Session: {st.session_state.session_id})")
        
        # --- TÍNH TOÁN SỐ NGƯỜI ONLINE (TRONG 5 PHÚT QUA) ---
        current_time = time.time()
        timeout_seconds = 300 # 5 phút = 300 giây
        
        # Đếm những ai vừa thao tác trong vòng 5 phút trở lại đây
        online_count = sum(1 for last_active in active_users.values() if current_time - last_active < timeout_seconds)
        
        # Khi đã đăng nhập thành công
        # st.success(f"Xin chào Admin! (Session: {st.session_state.session_id})")
        
        # --- LẤY DỮ LIỆU ĐẾM ---
        total_visits = get_total_access()
        
        # Đếm số người online trong 5 phút qua và dọn dẹp RAM
        current_time = time.time()
        timeout_seconds = 300 
        users_to_remove = [sid for sid, last_active in active_users.items() if current_time - last_active > timeout_seconds]
        for sid in users_to_remove:
            del active_users[sid]
        current_online = len(active_users)
        
        # --- THANH CÔNG CỤ ADMIN ---
        col_act1, col_act2, col_act3 = st.columns([1.5, 1.5, 1.5])
        with col_act1:
            st.metric(label="🟢 Đang Online", value=f"{current_online} người")
            if st.button("Đăng xuất", type="secondary"):
                st.session_state.admin_logged_in = False
                st.rerun()
                
        with col_act2:
            st.metric(label="📊 Tổng lượt truy cập", value=f"{total_visits} lượt")
            
        with col_act3:
            st.markdown("<div style='margin-top: 35px;'></div>", unsafe_allow_html=True) # Căn lề cho đẹp
            # Checkbox tự động refresh vẫn giữ nguyên
            auto_refresh = st.checkbox("🔄 Tự động làm mới (30s)", value=st.session_state.auto_refresh)
            st.session_state.auto_refresh = auto_refresh
            
        # --- TÍNH NĂNG DEBUG HEADERS ĐỂ TÌM IP ---
        with st.expander("🕵️ Debug: Xem Headers (Tìm IP thật)"):
            if hasattr(st, "context") and st.context.headers:
                st.json(dict(st.context.headers))
                st.caption("Nếu bạn thấy IP thật nằm trong trường nào, hãy báo Dev để cấu hình lại.")
            else:
                st.warning("Không tìm thấy Header nào.")

        st.markdown("---")
        st.subheader("📜 Nhật ký hệ thống (System Logs)")
        
        # Hiển thị log
        if os.path.exists(LOG_FILE):
            try:
                with open(LOG_FILE, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                    preview_lines = "".join(lines[-200:]) # 200 dòng mới nhất
                
                # Khu vực hiển thị log
                st.code(preview_lines, language="log")
                
                with open(LOG_FILE, "rb") as f:
                    st.download_button(
                        label="📥 Tải toàn bộ file Log",
                        data=f,
                        file_name="pmr_tool_usage.log",
                        mime="text/plain"
                    )
            except Exception as e:
                st.error(f"Lỗi đọc file log: {e}")
        else:
            st.warning("Chưa có file log nào được tạo.")

        # Xử lý Logic Auto Refresh
        if st.session_state.auto_refresh:
            time.sleep(30)
            st.rerun()
            
        st.markdown("---")
        st.caption(f"Phiên bản App: {APP_VERSION}")

else:
    # ---------------------------------------------------------
    # GIAO DIỆN NGƯỜI DÙNG BÌNH THƯỜNG (TOOL CHÍNH)
    # ---------------------------------------------------------
    
    # Hạn chế kích thước upload
    MAX_UPLOAD_MB = 50
    MAX_UPLOAD_BYTES = MAX_UPLOAD_MB * 1024 * 1024

    banner_file = "logo_CTS.jpg" 
    if os.path.exists(banner_file):
        st.image(banner_file)
    else:
        pass # Xóa thông báo warning để đỡ rối giao diện

    # Nội dung hướng dẫn
    help_html = """<span class='tooltip-container'>📖 Hướng dẫn sử dụng
    <span class='tooltiptext'>
    <h4 style='margin:0; text-align:center; color:#0068C9'>HƯỚNG DẪN SỬ DỤNG NHANH</h4><hr>
    <strong>1. Chuẩn bị dữ liệu đầu vào</strong><br>
    • File Excel (.xlsx) xuất từ phần mềm cấp phép ( phiên bản windows) chứa các trạm hiện hữu. Lưu ý dữ liệu xuất phải là dữ liệu mới nhất có thể<br>
    • Cho phép nạp tối đã 02 files<br>
    • Các Cột cần có: Số GP, Tần số, Tọa độ, Độ cao, Khách hàng, tỉnh thành.<br>
    <strong>2. Nhập thông số (Cột bên trái)</strong><br>
    • Nhập Tọa độ, Loại mạng (LAN/WAN), Độ cao, Dải tần.<br>
    • Tiện ích Convert tọa độ cho phép nhập dưới dạng thập phân và tự động điền vào dưới dạng độ,phút giây.<br>
    • <strong>Đoạn băng tần quét:</strong> Chọn dải tần con (VD: 141.5 - 142.0).<br>
    • Chọn Tỉnh/TP (với mạng LAN).<br><br>
    <strong>3. Các chức năng tính toán</strong><br>
    • <strong>TÍNH TẦN SỐ KHẢ DỤNG:</strong> Tìm tần số sạch, sắp xếp theo độ ưu tiên.<br>
    • <strong>KIỂM TRA CỤ THỂ:</strong> Kiểm tra nhanh 1 tần số bất kỳ xem có khả dụng không. Cần nhập tọa độ, dải tần, loại mạng, độ cao ( LAN), băng thông và tần số cần kiểm tra<br><br>
    <strong>4. Lưu kết quả</strong><br>
    • Nút <strong>📥 LƯU KẾT QUẢ (EXCEL)</strong> sẽ xuất hiện sau khi tính xong.<br><br>
    <strong>5. Cách đọc kết quả</strong><br>
    • Cột "Các GP sử dụng tần số này" hiển thị dạng: <strong>SốGP(Khoảng_cách_km)</strong><br>
    • VD: 399123(65) -> GP 399123 đang dùng, cách 65km.
    </span>
    </span>"""

    st.markdown("<h2 style='text-align: center; color: #0068C9;'>Ấn định tần số cho mạng nội bộ dùng riêng </h2>", unsafe_allow_html=True)
    st.markdown(f"<div style='text-align: right; color: #666; font-size:0.85rem; margin-top:-8px;'>{help_html} | Phiên bản: {APP_VERSION}</div>", unsafe_allow_html=True)
    st.markdown("---")

    col_layout_left, col_space_layout, col_layout_right = st.columns([1.8, 0.1, 1.2])

    with col_layout_left:
        st.subheader("1. THÔNG SỐ KỸ THUẬT & VỊ TRÍ MẠNG")
        c_grp1, c_sep1, c_grp2, c_sep2, c_grp3 = st.columns([1.3, 0.1, 1.3, 0.1, 1.5])
        
        with c_grp1:
            st.markdown("📍 **Kinh độ (Longitude)**")
            c1_d, c1_m, c1_s = st.columns([1, 1, 1.2])
            with c1_d: lon_d = st.number_input("Độ", min_value=0, max_value=180, step=1, key="lon_d", label_visibility="collapsed")
            with c1_m: lon_m = st.number_input("Phút", min_value=0, max_value=59, step=1, key="lon_m", label_visibility="collapsed")
            with c1_s: lon_s = st.number_input("Giây", min_value=0.0, max_value=59.99, step=0.1, format="%.2f", key="lon_s", label_visibility="collapsed")
            lon = dms_to_decimal(lon_d, lon_m, lon_s)

        with c_grp2:
            st.markdown("📍 **Vĩ độ (Latitude)**")
            c2_d, c2_m, c2_s = st.columns([1, 1, 1.2])
            with c2_d: lat_d = st.number_input("Độ", min_value=0, max_value=90, step=1, key="lat_d", label_visibility="collapsed")
            with c2_m: lat_m = st.number_input("Phút", min_value=0, max_value=59, step=1, key="lat_m", label_visibility="collapsed")
            with c2_s: lat_s = st.number_input("Giây", min_value=0.0, max_value=59.99, step=0.1, format="%.2f", key="lat_s", label_visibility="collapsed")
            lat = dms_to_decimal(lat_d, lat_m, lat_s)

        with c_grp3:
            st.markdown("🗺️ **Bản đồ**")
            col_conv, col_map = st.columns(2)
            with col_conv:
                if st.button("Tọa độ Decimal", use_container_width=True): show_convert_popup()
            with col_map:
                if lat != 0 and lon != 0:
                    if st.button("Vị trí trên map", use_container_width=True): show_map_popup(lat, lon)
                else: st.button("Vị trí trên map", disabled=True, use_container_width=True)

        c_mode, c_h, c_band, c_subband, c_bw = st.columns([1.2, 0.7, 0.7, 1.6, 0.8], gap="small")
        
        with c_mode:
            st.markdown("📡 **Loại mạng**")
            mode = st.selectbox("Loại mạng", ["LAN", "WAN_SIMPLEX", "WAN_DUPLEX"], label_visibility="collapsed")

        with c_h:
            st.markdown("**Độ cao (m)**")
            h_anten = st.number_input("Độ cao", value=0.0, step=1.0, label_visibility="collapsed")
        
        with c_band:
            st.markdown("**Dải tần**")
            band = st.selectbox("Dải tần", ["VHF", "UHF"], label_visibility="collapsed")
            
        with c_subband:
            st.markdown("**Đoạn băng tần quét**")
            if band == "VHF":
                current_alloc = ALLOC_VHF
            else:
                current_alloc = ALLOC_UHF
                
            subband_map = {}
            subband_labels = []
            
            for item in current_alloc:
                s_f, e_f, m_list, note = item
                if mode in m_list:
                    label = f"{s_f} - {e_f} MHz ({note})"
                    subband_map[label] = (s_f, e_f)
                    subband_labels.append(label)
                
            selected_subband_label = st.selectbox("Chọn dải con", subband_labels, label_visibility="collapsed")
            scan_start, scan_end = subband_map.get(selected_subband_label, (0, 0))

        with c_bw:
            st.markdown("**Băng thông**")
            bw = st.selectbox("Băng thông", [6.25, 12.5, 25.0], index=1, label_visibility="collapsed")
        
        c_prov, c_qty, c_space = st.columns([1.2, 0.8, 3.0], gap="small")
        with c_prov:
            st.markdown("**Tỉnh / Thành phố**")
            is_wan = "WAN" in mode
            province_selection = st.selectbox("Chọn Tỉnh/TP", ["-- Chọn Tỉnh/TP --", "HANOI", "HCM", "DANANG", "KHAC"], index=0, label_visibility="collapsed", disabled=is_wan)
        
        with c_qty:
            st.markdown("**Số lượng tần số**")
            qty = st.number_input("Số lượng", value=1, min_value=1, label_visibility="collapsed")
        
        with c_space:
            st.empty() 

    with col_layout_right:
        st.subheader("2. NẠP DỮ LIỆU ĐẦU VÀO")
        uploaded_files = st.file_uploader("Label ẩn", type=None, label_visibility="collapsed", accept_multiple_files=True)
        
        btn_disabled = True 
        if uploaded_files:
            if len(uploaded_files) > 2:
                st.error("⚠️ Bạn chỉ được phép nạp tối đa 2 files cùng lúc.")
                btn_disabled = True
            else:
                total_size = sum(getattr(f, "size", 0) for f in uploaded_files)
                if total_size > MAX_UPLOAD_BYTES:
                    st.error(f"Tổng dung lượng files quá lớn (> {MAX_UPLOAD_MB} MB).")
                    btn_disabled = True
                elif not all(f.name.lower().endswith('.xlsx') for f in uploaded_files):
                    st.error("⚠️ Tất cả các file nạp vào phải có định dạng .xlsx")
                    btn_disabled = True
                else:
                    current_file_id = "_".join([f"{f.name}_{getattr(f, 'size', '')}" for f in uploaded_files])
                    
                    if st.session_state.last_uploaded_file_id != current_file_id:
                        st.cache_resource.clear() 
                        st.session_state.results = None
                        st.session_state.input_snapshot = None
                        st.session_state.check_result = None
                        st.session_state.bad_freq_results = None
                        st.session_state.active_view = None
                        st.session_state.last_uploaded_file_id = current_file_id
                        
                        file_names = ", ".join([f.name for f in uploaded_files])
                        log_info(f"SESS: {st.session_state.session_id} | ACTION: UPLOAD | Files: {file_names} | Total Size: {total_size}")
                        st.rerun() 
                    
                    try:
                        get_tool_instance(uploaded_files)
                        btn_disabled = False 
                    except ValueError as ve:
                        st.error(f"❌ Lỗi dữ liệu đầu vào: {ve}")
                        btn_disabled = True
                    except Exception as e:
                        st.error(f"❌ Lỗi hệ thống: {e}")
                        btn_disabled = True
                        
        else:
            if st.session_state.last_uploaded_file_id is not None:
                st.cache_resource.clear() 
                st.session_state.results = None
                st.session_state.input_snapshot = None
                st.session_state.check_result = None
                st.session_state.bad_freq_results = None
                st.session_state.active_view = None
                st.session_state.last_uploaded_file_id = None
                st.rerun()
        
        st.markdown('<div style="margin-top: -25px;"></div>', unsafe_allow_html=True)
        c_btn1, c_btn2 = st.columns(2)
        with c_btn1:
            btn_calc = st.button("TÍNH TẦN SỐ KHẢ DỤNG", type="primary", use_container_width=True, disabled=btn_disabled)
        with c_btn2:
            btn_scan_bad_freq = False  # Gán mặc định bằng False để vô hiệu hóa chức năng này

    st.markdown("---")
    st.subheader("3. KIỂM TRA TẦN SỐ CỤ THỂ")
    c_check_1, c_check_2 = st.columns([1.0, 4.0]) 
    with c_check_1:
        f_check_val = st.number_input("Nhập tần số (MHz):", value=0.0, step=0.0125, format="%.4f")
    with c_check_2:
        st.markdown(" ") 
        st.markdown(" ")
        btn_check_specific = st.button("KIỂM TRA CAN NHIỄU", type="secondary", disabled=btn_disabled)

    # =========================================================================
    # XỬ LÝ LOGIC (CHỈ CHẠY KHI Ở GIAO DIỆN CHÍNH)
    # =========================================================================

    if btn_calc:
        st.session_state.check_result = None
        st.session_state.bad_freq_results = None
        st.session_state.active_view = "AVAILABLE"
        
        error_msg = []
        if lon == 0.0: error_msg.append("Kinh độ chưa nhập")
        if lat == 0.0: error_msg.append("Vĩ độ chưa nhập")
        if "LAN" in mode:
            if province_selection == "-- Chọn Tỉnh/TP --": error_msg.append("Thiếu Tỉnh/TP (Bắt buộc cho mạng LAN)")
        
        if error_msg:
            st.error(f"⚠️ LỖI: {', '.join(error_msg)}")
            st.session_state.active_view = None
            log_warning(f"SESS: {st.session_state.session_id} | ACTION: CALC_ERROR | Msg: {', '.join(error_msg)}")
        else:
            prov_to_send = province_selection
            if "WAN" in mode: prov_to_send = "KHAC"
            if h_anten == 0.0: st.warning("⚠️ Lưu ý: Độ cao Anten đang là 0m.")
            
            log_info(f"SESS: {st.session_state.session_id} |CALC_START|Pos:{lat:.6f},{lon:.6f}|Mode:{mode}|Band:{band}|H:{h_anten}|Subband:{selected_subband_label}|Prov:{prov_to_send}")
            with st.spinner('Đang tính toán...'):
                try:
                    # --- SỬ DỤNG HÀM CACHE LẤY INSTANCE THAY VÌ TẠO MỚI ---
                    tool = get_tool_instance(uploaded_files)
                    
                    user_input = {
                        "lat": lat, "lon": lon,
                        "province_code": prov_to_send,
                        "antenna_height": h_anten,
                        "band": band, "bw": bw, "usage_mode": mode,
                        "scan_start": scan_start, "scan_end": scan_end 
                    }
                    results = tool.tinh_toan(user_input)
                    st.session_state.results = results
                    st.session_state.input_snapshot = {
                        "THAM SỐ": ["Phiên bản App", "Kinh độ", "Vĩ độ", "Tỉnh / TP", "Độ cao Anten (m)", "Dải tần", "Phạm vi quét", "Băng thông", "Loại mạng", "Số lượng xin"],
                        "GIÁ TRỊ": [APP_VERSION, f"{lon:.5f}", f"{lat:.5f}", prov_to_send if "LAN" in mode else "Toàn quốc (WAN)", h_anten, band, selected_subband_label, bw, mode, qty]
                    }
                    log_info(f"SESS: {st.session_state.session_id} | ACTION: CALC_SUCCESS | Found: {len(results)} freqs")
                    
                    # --- ÉP GIẢI PHÓNG RAM NGAY SAU KHI TÍNH TOÁN XONG ---
                    gc.collect()

                except Exception as e:
                    log_exception(f"SESS: {st.session_state.session_id} | ACTION: CALC_EXCEPTION | Error: {e}")
                    st.error(f"Có lỗi xảy ra: {e}")
                    st.session_state.active_view = None

    if btn_scan_bad_freq:
        st.session_state.results = None
        st.session_state.check_result = None
        st.session_state.active_view = "UNAVAILABLE"
        
        if not uploaded_files:
            st.error("Vui lòng nạp file Excel trước.")
            st.session_state.active_view = None
        elif btn_disabled: 
             st.error("Vui lòng nạp đúng định dạng file (.xlsx).")
             st.session_state.active_view = None
        else:
            prov_to_send = province_selection
            if "WAN" in mode: prov_to_send = "KHAC"
            
            log_info(f"SESS: {st.session_state.session_id} | ACTION: SCAN_BAD_START | Pos: {lat:.6f},{lon:.6f} | Mode: {mode} | Band: {band} | Subband: {selected_subband_label}")
            with st.spinner("Đang quét dải tần đã chọn..."):
                try:
                    # --- SỬ DỤNG HÀM CACHE ---
                    tool = get_tool_instance(uploaded_files)
                    
                    user_input = {
                        "lat": lat, "lon": lon,
                        "province_code": prov_to_send,
                        "antenna_height": h_anten,
                        "band": band, "bw": bw, "usage_mode": mode,
                        "scan_start": scan_start, "scan_end": scan_end 
                    }
                    bad_results = tool.tim_cac_tan_so_khong_kha_dung(user_input)
                    st.session_state.bad_freq_results = bad_results
                    
                    st.session_state.input_snapshot = {
                        "THAM SỐ": ["Phiên bản App", "Kinh độ", "Vĩ độ", "Tỉnh / TP", "Độ cao Anten (m)", "Dải tần", "Phạm vi quét", "Băng thông", "Loại mạng", "Số lượng xin"],
                        "GIÁ TRỊ": [APP_VERSION, f"{lon:.5f}", f"{lat:.5f}", prov_to_send if "LAN" in mode else "Toàn quốc (WAN)", h_anten, band, selected_subband_label, bw, mode, qty]
                    }
                    
                    log_info(f"SESS: {st.session_state.session_id} | ACTION: SCAN_BAD_SUCCESS | Found: {len(bad_results)} bad freqs")
                    
                    # --- GIẢI PHÓNG RAM ---
                    gc.collect()

                except Exception as e:
                    log_exception(f"SESS: {st.session_state.session_id} | ACTION: SCAN_BAD_EXCEPTION | Error: {e}")
                    st.error(f"Có lỗi xảy ra: {e}")
                    st.session_state.active_view = None

    if btn_check_specific:
        st.session_state.results = None
        st.session_state.bad_freq_results = None
        st.session_state.active_view = "CHECK_SPECIFIC"

        if not uploaded_files:
            st.error("Vui lòng nạp file Excel trước.")
            st.session_state.active_view = None
        elif btn_disabled:
             st.error("Vui lòng nạp đúng định dạng file (.xlsx).")
             st.session_state.active_view = None
        elif f_check_val <= 0:
            st.error("Vui lòng nhập tần số hợp lệ.")
            st.session_state.active_view = None
            log_warning(f"SESS: {st.session_state.session_id} | ACTION: CHECK_ERROR | Invalid Freq: {f_check_val}")
        else:
            prov_to_send = province_selection
            if "WAN" in mode: prov_to_send = "KHAC"
            
            log_info(f"SESS: {st.session_state.session_id} | CHECK_START | Freq: {f_check_val} | Pos: {lat:.6f},{lon:.6f}")
            with st.spinner(f"Đang kiểm tra tần số {f_check_val} MHz..."):
                try:
                    # --- SỬ DỤNG HÀM CACHE ---
                    tool = get_tool_instance(uploaded_files)
                    
                    user_input = {
                        "lat": lat, "lon": lon,
                        "province_code": prov_to_send,
                        "antenna_height": h_anten,
                        "band": band, "bw": bw, "usage_mode": mode
                    }
                    check_res = tool.kiem_tra_tan_so_cu_the(user_input, f_check_val)
                    st.session_state.check_result = check_res
                    
                    status = check_res.get("status", "UNKNOWN")
                    log_info(f"SESS: {st.session_state.session_id} | ACTION: CHECK_SUCCESS | Status: {status}")
                    
                    # --- GIẢI PHÓNG RAM ---
                    gc.collect()

                except Exception as e:
                    log_exception(f"SESS: {st.session_state.session_id} | ACTION: CHECK_EXCEPTION | Error: {e}")
                    st.error(f"Có lỗi xảy ra: {e}")
                    st.session_state.active_view = None

    # VIEW 1: KẾT QUẢ TẦN SỐ KHẢ DỤNG
    if st.session_state.active_view == "AVAILABLE" and st.session_state.results is not None:
        st.markdown("---")
        st.subheader("📊 KẾT QUẢ TÍNH TOÁN: TẦN SỐ KHẢ DỤNG")
        results = st.session_state.results
        
        if not results:
            st.error("❌ Không tìm thấy tần số khả dụng trong dải quét!")
        else:
            df_res = pd.DataFrame(results)
            cols_display = ["STT", "frequency", "reuse_factor", "license_list"]
            df_view = df_res[cols_display].copy()
            df_view.columns = ["STT", "Tần số Khả dụng (MHz)", "Hệ số Tái sử dụng (Điểm)", "Các GP sử dụng tần số này"]
            df_view.set_index("STT", inplace=True)

            m1, m2 = st.columns(2)
            m1.metric("Số lượng tìm thấy", f"{len(results)}")
            best_freq = results[0]['frequency']
            m2.metric("Tần số tốt nhất", f"{best_freq} MHz")

            df_top = df_view.head(qty)

            def style_logic(df):
                styles = pd.DataFrame('', index=df.index, columns=df.columns)
                for idx in df.index:
                    row_data = df_res[df_res['STT'] == idx].iloc[0]
                    is_prio = row_data.get('is_priority', False)
                    if is_prio:
                        styles.loc[idx, :] = f'color: {PRIORITY_HIGHLIGHT_COLOR}; font-weight: bold'
                    elif idx <= results[min(qty-1, len(results)-1)]['STT']: 
                        top_ids = [item['STT'] for item in results[:qty]]
                        if idx in top_ids:
                            styles.loc[idx, :] = 'color: #28a745; font-weight: bold'
                return styles

            styler_top = df_top.style.apply(lambda x: style_logic(df_top), axis=None)
            styler_full = df_view.style.apply(lambda x: style_logic(df_view), axis=None)

            st.markdown(f"**Danh sách {qty} tần số đề xuất tốt nhất:**")
            st.table(styler_top)
            
            with st.expander("Xem danh sách đầy đủ (Tất cả kết quả)"):
                st.dataframe(styler_full, use_container_width=True)

            if st.session_state.input_snapshot:
                df_export = df_res.copy()
                df_export.rename(columns={
                    "STT": "STT",
                    "frequency": "Tần số Khả dụng (MHz)",
                    "reuse_factor": "Hệ số Tái sử dụng",
                    "license_list": "Các GP sử dụng tần số này (kèm khoảng cách)"
                }, inplace=True)
                
                df_input_report = pd.DataFrame(st.session_state.input_snapshot)
                excel_data = to_excel(df_input_report, df_export)
                
                now = datetime.now()
                time_str = now.strftime("%H%M%S_%d%m%Y")
                
                input_file_name = "data"
                if uploaded_files:
                    # Lấy tên của 1-2 file đầu tiên ghép lại để tải về
                    input_file_name = "_".join([os.path.splitext(f.name)[0] for f in uploaded_files][:2])
                    
                dl_file_name = f"DS_TanSo_KhaDung_{time_str}_{input_file_name}.xlsx"
                
                st.markdown("---")
                st.download_button(
                    label=f"📥 LƯU KẾT QUẢ (EXCEL)",
                    data=excel_data,
                    file_name=dl_file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    # VIEW 2: KẾT QUẢ TẦN SỐ KHÔNG KHẢ DỤNG
    elif st.session_state.active_view == "UNAVAILABLE" and st.session_state.bad_freq_results is not None:
        st.markdown("---")
        st.subheader("⚠️ CÁC TẦN SỐ KHÔNG KHẢ DỤNG (GÂY NHIỄU)")
        
        bad_list = st.session_state.bad_freq_results
        if not bad_list:
            st.info("Tuyệt vời! Không tìm thấy tần số nào bị nhiễu trong dải quét.")
        else:
            st.warning(f"⚠️ Tìm thấy {len(bad_list)} trường hợp tần số gây nhiễu (không khả dụng).")
            df_bad = pd.DataFrame(bad_list)
            
            if "Khoảng cách thực tế (km)" in df_bad.columns and "Khoảng cách yêu cầu (km)" in df_bad.columns:
                 df_bad["Khoảng cách thực tế/Chỉ tiêu"] = df_bad.apply(lambda x: f"{x['Khoảng cách thực tế (km)']:.2f}/{x['Khoảng cách yêu cầu (km)']:.2f}", axis=1)
                 df_bad.drop(columns=["Khoảng cách thực tế (km)", "Khoảng cách yêu cầu (km)"], inplace=True)
            elif "dist_km" in df_bad.columns and "req_dist_km" in df_bad.columns:
                 df_bad["Khoảng cách thực tế/Chỉ tiêu"] = df_bad.apply(lambda x: f"{x['dist_km']:.2f}/{x['req_dist_km']:.2f}", axis=1)
                 df_bad.drop(columns=["dist_km", "req_dist_km"], inplace=True)
                 
            st.dataframe(
                df_bad, 
                use_container_width=True,
                column_config={
                    "Tên Khách Hàng": st.column_config.TextColumn(width="large"), 
                    "Địa chỉ trạm bị nhiễu": st.column_config.TextColumn(width="medium"),
                    "Khoảng cách thực tế/Chỉ tiêu": st.column_config.TextColumn(width="medium", label="K.Cách Thực tế/Chỉ tiêu (km)"),
                    "Tần số (MHz)": st.column_config.NumberColumn(format="%.4f"),
                    "Tần số trạm bị nhiễu (MHz)": st.column_config.NumberColumn(format="%.4f"),
                }
            )
            
            if st.session_state.input_snapshot:
                df_input_report = pd.DataFrame(st.session_state.input_snapshot)
                csv_data = to_excel(df_input_report, df_bad)
            else:
                csv_data = to_excel(None, df_bad)
                
            now = datetime.now()
            time_str = now.strftime("%H%M%S_%d%m%Y")
            
            input_file_name = "data"
            if uploaded_files:
                input_file_name = "_".join([os.path.splitext(f.name)[0] for f in uploaded_files][:2])
                
            dl_file_name = f"DS_TanSo_KhongKhaDung_{time_str}_{input_file_name}.xlsx"
            
            st.download_button(
                label="📥 Tải danh sách Excel",
                data=csv_data,
                file_name=dl_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    # VIEW 3: KẾT QUẢ KIỂM TRA CỤ THỂ
    elif st.session_state.active_view == "CHECK_SPECIFIC" and st.session_state.check_result is not None:
        st.markdown("---")
        st.subheader("🔎 KẾT QUẢ KIỂM TRA TẦN SỐ CỤ THỂ")
        
        res = st.session_state.check_result
        if res.get("status") == "OK":
            st.success(f"✅ {res.get('msg')}")
        else:
            st.error(f"❌ {res.get('msg')}")
            if "conflicts" in res and res["conflicts"]:
                st.markdown("**Danh sách các giấy phép gây nhiễu (không đảm bảo khoảng cách):**")
                df_conflict = pd.DataFrame(res["conflicts"])
                if not df_conflict.empty:
                    df_conflict["Khoảng cách thực tế/Chỉ tiêu"] = df_conflict.apply(lambda x: f"{x['dist_km']:.2f}/{x['req_dist_km']:.2f}", axis=1)
                    df_conflict.drop(columns=["dist_km", "req_dist_km"], inplace=True)
                    
                    df_conflict.rename(columns={
                        "license": "Số Giấy Phép",
                        "customer": "Tên Khách Hàng",
                        "freq_conflict": "Tần số GP (MHz)",
                        "address": "Địa chỉ trạm",
                        "type": "Loại nhiễu"
                    }, inplace=True)

                    st.table(df_conflict)

# =============================================================================
# --- LUÔN GIẢI PHÓNG BỘ NHỚ SAU MỖI LẦN RENDER TRANG CHỐNG TRÀN RAM ---
gc.collect()
# =============================================================================